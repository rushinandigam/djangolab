import tkinter as tk
from tkinter import messagebox
import csv
import openpyxl

# Function to check for duplicates
def check_duplicates(student_id, email_id):
    try:
        with open('students.csv', mode='r') as file:
            reader = csv.reader(file)
            for row in reader:
                if row[1] == student_id or row[2] == email_id:
                    return True
    except FileNotFoundError:
        pass
    
    try:
        workbook = openpyxl.load_workbook('students.xlsx')
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            if row[1] == student_id or row[2] == email_id:
                return True
    except FileNotFoundError:
        pass
    
    return False

# Function to handle the submission
def submit_form():
    student_name = entry_name.get()
    student_id = entry_id.get()
    email_id = entry_email.get()
    phone_number = entry_phone.get()
    address = entry_address.get()

    if not student_name or not student_id or not email_id or not phone_number or not address:
        messagebox.showerror("Error", "Please fill in all fields!")
    elif check_duplicates(student_id, email_id):
        messagebox.showerror("Error", "Record already exists!")
    else:
        # If all fields are filled and no duplicates, show a success message
        messagebox.showinfo("Success", "Form submitted successfully!")
        
        # Write data to CSV file
        with open('students.csv', mode='a', newline='') as file:
            writer = csv.writer(file)
            writer.writerow([student_name, student_id, email_id, phone_number, address])
        
        # Write data to Excel file
        try:
            workbook = openpyxl.load_workbook('students.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Student Name", "Student ID", "Email ID", "Phone Number", "Address"])
        
        sheet.append([student_name, student_id, email_id, phone_number, address])
        workbook.save('students.xlsx')

# Function to delete a record
def delete_record():
    student_id = entry_delete_id.get()
    if not student_id:
        messagebox.showerror("Error", "Please enter a Student ID!")
        return

    # Delete from CSV file
    try:
        with open('students.csv', mode='r') as file:
            rows = list(csv.reader(file))
        with open('students.csv', mode='w', newline='') as file:
            writer = csv.writer(file)
            record_found = False
            for row in rows:
                if row[1] != student_id:
                    writer.writerow(row)
                else:
                    record_found = True
    except FileNotFoundError:
        record_found = False

    # Delete from Excel file
    try:
        workbook = openpyxl.load_workbook('students.xlsx')
        sheet = workbook.active
        record_found_excel = False
        for row in sheet.iter_rows(values_only=True):
            if row[1] == student_id:
                record_found_excel = True
                sheet.delete_rows(row[0].row)
                break
        workbook.save('students.xlsx')
    except FileNotFoundError:
        record_found_excel = False

    if record_found or record_found_excel:
        messagebox.showinfo("Success", "Record deleted successfully!")
    else:
        messagebox.showerror("Error", "Record not found!")

# Function to open the delete window
def open_delete_window():
    delete_window = tk.Toplevel(root)
    delete_window.title("Delete Record")
    delete_window.geometry("300x200")
    delete_window.config(bg="#FF6347")  # Tomato background

    label_delete_id = tk.Label(delete_window, text="Student ID:", bg="#FF6347", font=("Arial", 12))
    label_delete_id.pack(pady=10)
    global entry_delete_id
    entry_delete_id = tk.Entry(delete_window, font=("Arial", 12))
    entry_delete_id.pack(pady=10)

    delete_button = tk.Button(delete_window, text="Delete", command=delete_record, bg="#FF6347", font=("Arial", 12))
    delete_button.pack(pady=20)

# Function to clear the input fields
def clear_fields():
    entry_name.delete(0, tk.END)
    entry_id.delete(0, tk.END)
    entry_email.delete(0, tk.END)
    entry_phone.delete(0, tk.END)
    entry_address.delete(0, tk.END)

# Create the main window
root = tk.Tk()
root.title("Student Login Form")
root.geometry("400x400")
root.config(bg="#90EE90")  # Light green background

# Create a frame to center the fields
frame = tk.Frame(root, bg="#90EE90")
frame.place(relx=0.5, rely=0.5, anchor="center")

# Create the labels and entry fields
label_name = tk.Label(frame, text="Student Name:", bg="#90EE90", font=("Arial", 12))
label_name.grid(row=0, column=0, padx=10, pady=5, sticky="w")
entry_name = tk.Entry(frame, font=("Arial", 12))
entry_name.grid(row=0, column=1, padx=10, pady=5)

label_id = tk.Label(frame, text="Student ID:", bg="#90EE90", font=("Arial", 12))
label_id.grid(row=1, column=0, padx=10, pady=5, sticky="w")
entry_id = tk.Entry(frame, font=("Arial", 12))
entry_id.grid(row=1, column=1, padx=10, pady=5)

label_email = tk.Label(frame, text="Email ID:", bg="#90EE90", font=("Arial", 12))
label_email.grid(row=2, column=0, padx=10, pady=5, sticky="w")
entry_email = tk.Entry(frame, font=("Arial", 12))
entry_email.grid(row=2, column=1, padx=10, pady=5)

label_phone = tk.Label(frame, text="Phone Number:", bg="#90EE90", font=("Arial", 12))
label_phone.grid(row=3, column=0, padx=10, pady=5, sticky="w")
entry_phone = tk.Entry(frame, font=("Arial", 12))
entry_phone.grid(row=3, column=1, padx=10, pady=5)

label_address = tk.Label(frame, text="Address:", bg="#90EE90", font=("Arial", 12))
label_address.grid(row=4, column=0, padx=10, pady=5, sticky="w")
entry_address = tk.Entry(frame, font=("Arial", 12))
entry_address.grid(row=4, column=1, padx=10, pady=5)

# Submit button
submit_button = tk.Button(frame, text="Submit", command=submit_form, bg="#90EE90", font=("Arial", 12))
submit_button.grid(row=5, column=0, pady=20)

# Clear button
clear_button = tk.Button(frame, text="Clear", command=clear_fields, bg="#90EE90", font=("Arial", 12))
clear_button.grid(row=5, column=1, pady=20)

# Delete button to open the delete window at the right corner
delete_window_button = tk.Button(root, text="Delete Record", command=open_delete_window, bg="#FF6347", font=("Arial", 12))
delete_window_button.place(relx=0.95, rely=0.95, anchor="se")

# Start the Tkinter event loop
root.mainloop()
